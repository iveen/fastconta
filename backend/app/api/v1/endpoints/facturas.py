# app/api/v1/endpoints/facturas.py
import logging
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import List
from uuid import UUID

import xlrd
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.file_handlers import FileHandlerRegistry
from app.core.security import DataScope, get_data_scope
from app.db.session import get_public_db
from app.dependencies.empresa import get_active_empresa  # ✅ NUEVO
from app.models.global_models import CatalogoMoneda, TipoDTE
from app.models.tenant_models import (
    CuentaContable,
    DetallePartida,
    Empresa,  # ✅ NUEVO
    FacturaDetalle,
    FacturaElectronica,
    Partida,
)
from app.schemas.factura import FacturaOut
from app.schemas.partida import DetallePartidaOut, PartidaOut
from app.services.banguat_ws import obtener_tipo_cambio
from app.services.fel.context import FelIngestionContext

logger = logging.getLogger(__name__)
router = APIRouter()

# ==========================================
# Helper: Configurar search_path según rol
# ==========================================
async def _set_schema_for_query(db: AsyncSession, scope: DataScope, tenant_id: str | None = None):
    """Configura el search_path correcto según el rol del usuario."""
    schema_name = None
    if scope.role_code == "superadmin":
        if not tenant_id:
            raise HTTPException(400, detail="Superadmin debe especificar un tenant_id")
        res = await db.execute(
            text("SELECT schema_name FROM public.tenants WHERE id = :tid"),
            {"tid": tenant_id}
        )
        row = res.first()
        if not row:
            raise HTTPException(404, detail="Tenant no encontrado")
        schema_name = row[0]
    else:
        res = await db.execute(
            text("SELECT schema_name FROM public.tenants WHERE id = :tid"),
            {"tid": str(scope.tenant_id)}
        )
        row = res.first()
        if not row:
            raise HTTPException(404, detail="Tenant no encontrado")
        schema_name = row[0]

    if not schema_name.replace("_", "").isalnum():
        raise HTTPException(500, detail="Schema con formato inválido")

    await db.execute(text(f"SET LOCAL search_path TO {schema_name}, public"))
    return schema_name


# ==========================================
# Helper: Clasificación Automática de Gasto SAT
# ==========================================
async def clasificar_gasto_sat(datos: dict) -> str:
    """Clasifica el gasto basado en las descripciones de los items."""
    items = datos.get('items', [])
    descripcion_combinada = ' '.join([str(item.get('descripcion', '')).lower() for item in items])

    if 'gasolina' in descripcion_combinada or 'diesel' in descripcion_combinada or 'combustible' in descripcion_combinada:
        return 'COMBUSTIBLE'

    if 'medicamento' in descripcion_combinada or 'farmacia' in descripcion_combinada:
        return 'MEDICAMENTO'

    if any(p in descripcion_combinada for p in ['vehiculo', 'computadora', 'laptop', 'mobiliario', 'maquinaria', 'equipo']):
        return 'ACTIVO_FIJO'

    return 'NORMAL'


# ==========================================
# 1. Upload de Facturas XML
# ==========================================
@router.post("/upload", response_model=dict, status_code=status.HTTP_201_CREATED)
async def upload_facturas(
    empresa_id: str | None = Query(None, description="ID de la empresa (opcional, usa X-Company-Id)"),  # ✅ Cambiado a opcional
    tenant_id: str | None = Query(None, description="ID del tenant"),
    files: List[UploadFile] = File(...),
    scope: DataScope = Depends(get_data_scope),
    db: AsyncSession = Depends(get_public_db),
    empresa_from_header: Empresa | None = Depends(get_active_empresa)  # ✅ NUEVO
):
    await _set_schema_for_query(db, scope, tenant_id)
    
    # ✅ Usar empresa_id del header si no se pasó como query param
    empresa_id_final = empresa_id or (str(empresa_from_header.id) if empresa_from_header else None)
    
    if not empresa_id_final:
        raise HTTPException(400, detail="Debe especificar una empresa (query param o header X-Company-Id)")
    
    # Verificar empresa
    emp = await db.execute(text("SELECT id FROM empresas WHERE id = :eid"), {"eid": empresa_id_final})
    if not emp.first():
        raise HTTPException(400, "Empresa no encontrada")

    # Obtener NIT de la empresa actual para determinar si es Venta o Compra
    emp_nit_res = await db.execute(text("SELECT nit FROM empresas WHERE id = :eid"), {"eid": empresa_id_final})
    empresa_nit = emp_nit_res.scalar_one().replace('-', '').strip()

    facturas_creadas = []
    rechazos = []
    requieren_revision_manual = []

    for file in files:
        try:
            # 1. Leer con handler genérico
            handler = FileHandlerRegistry.resolve(file.filename, file.content_type)
            content = await handler.read(file)
            logger.info(f"📄 Archivo leído: {file.filename}, extensión: {content.extension}")
            
            # 2. Parsear con estrategia FEL
            result = await FelIngestionContext.ingest(content, db)
            logger.info(f"🔍 Resultado del parse: success={result.success}, error={result.error}, review={result.requires_manual_review}")
            
            if not result.success:
                logger.warning(f"❌ Parse falló: {result.error}")
                rechazos.append(f"{file.filename}: {result.error}")
                continue

            if result.requires_manual_review:
                requieren_revision_manual.append(file.filename)
                logger.info(f"⚠️ {file.filename} requiere revisión manual, pero se guardará")
            
            datos = result.data
            logger.info(f"✅ Datos extraídos: {datos.get('numero_autorizacion')}, total={datos.get('total')}")

            # Evitar duplicados
            dup = await db.execute(
                text("SELECT id FROM facturas_electronicas WHERE empresa_id = :e AND serie = :s AND numero_autorizacion = :n"),
                {"e": empresa_id_final, "s": datos.get('serie'), "n": datos.get('numero_autorizacion')}
            )
            if dup.first():
                rechazos.append(f"{file.filename}: Duplicada")
                continue

            # Determinar tipo de operación
            em = datos.get('emisor_nit','').replace('-','')
            rec = datos.get('receptor_nit','').replace('-','')
            
            if em == empresa_nit:
                tipo_op = 'Venta'
            elif rec == empresa_nit:
                tipo_op = 'Compra'
            else:
                logger.warning(
                    f"Factura {datos.get('numero_autorizacion')} rechazada: "
                    f"La empresa {empresa_nit} no es ni emisor ({em}) ni receptor ({rec})"
                )
                rechazos.append(f"{file.filename}: La empresa no participa en esta transacción")
                continue

            # Obtener IDs de catálogos
            res_tipo = await db.execute(select(TipoDTE.id).where(TipoDTE.codigo == datos.get('tipo_documento','FACT')))
            tipo_id = res_tipo.scalar_one_or_none()
            
            res_mon = await db.execute(select(CatalogoMoneda.id).where(CatalogoMoneda.codigo_iso == datos.get('moneda','GTQ')))
            mon_id = res_mon.scalar_one_or_none()

            tc = Decimal("1.00000")
            if datos.get('moneda') != 'GTQ' and datos.get('fecha_emision'):
                tc = await obtener_tipo_cambio(datos['fecha_emision'].date(), datos['moneda'], db) or tc

            items = datos.pop('items', [])

            # Calcular montos en GTQ
            tc = float(tc)
            total_gravado_gtq = float(datos['total_gravado']) * tc
            total_iva_gtq = float(datos['total_iva']) * tc
            total_exento_gtq = float(datos.get('total_exento', 0)) * tc
            total_gtq = float(datos['total']) * tc
            
            clasificacion_inicial = await clasificar_gasto_sat(datos)

            # ============================================
            # XML original (extraerlo del FileContent)
            # ============================================
            xml_original = content.parsed_data.get("xml_text", "") if content.parsed_data else None

            factura = FacturaElectronica(
                empresa_id=empresa_id_final,
                xml_original=xml_original,
                numero_autorizacion=datos['numero_autorizacion'],
                serie=datos.get('serie'),
                numero=datos.get('numero'),
                fecha_emision=datos['fecha_emision'],
                emisor_nit=datos['emisor_nit'],
                emisor_nombre=datos['emisor_nombre'],
                receptor_nit=datos['receptor_nit'],
                receptor_nombre=datos['receptor_nombre'],
                total_gravado=datos['total_gravado'],
                total_iva=datos['total_iva'],
                total_exento=datos.get('total_exento', 0),
                total=datos['total'],
                tipo_documento_id=tipo_id,
                moneda_id=mon_id,
                tipo_cambio=tc,
                pais_destino_exportacion=datos.get('pais_destino_exportacion'),
                total_gravado_gtq=total_gravado_gtq,
                total_iva_gtq=total_iva_gtq,
                total_exento_gtq=total_exento_gtq,
                total_gtq=total_gtq,
                total_gravado_bienes=datos.get('total_gravado_bienes', 0),
                total_iva_bienes=datos.get('total_iva_bienes', 0),
                total_gravado_servicios=datos.get('total_gravado_servicios', 0),
                total_iva_servicios=datos.get('total_iva_servicios', 0),
                total_gravado_bienes_gtq=datos.get('total_gravado_bienes_gtq', 0),
                total_iva_bienes_gtq=datos.get('total_iva_bienes_gtq', 0),
                total_gravado_servicios_gtq=datos.get('total_gravado_servicios_gtq', 0),
                total_iva_servicios_gtq=datos.get('total_iva_servicios_gtq', 0),
                es_exportacion=datos.get('es_exportacion', False),
                tipo_operacion=tipo_op,
                estado='Activa',
                tipo_documento=datos.get('tipo_documento'),
                moneda=datos.get('moneda'),
                xml_filename=file.filename,
                retencion_iva=Decimal(str(datos.get('retencion_iva', 0.0) or 0.0)),
                retencion_isr=Decimal(str(datos.get('retencion_isr', 0.0) or 0.0)),
                clasificacion_gasto_sat=clasificacion_inicial,
                es_importacion=bool(datos.get('es_importacion', False)),
                requiere_revision_manual=result.requires_manual_review,
            )
            db.add(factura)
            await db.flush()

            for it in items:
                precio_unitario_gtq = float(it['precio_unitario']) * tc
                total_linea_gtq = float(it['total_linea']) * tc
                iva_linea_gtq = float(it.get('iva_linea', 0)) * tc

                db.add(FacturaDetalle(
                    factura_id=factura.id,
                    cantidad=it['cantidad'],
                    descripcion=it['descripcion'],
                    precio_unitario=it['precio_unitario'],
                    total_linea=it['total_linea'],
                    iva_linea=it.get('iva_linea', 0),
                    precio_unitario_gtq=precio_unitario_gtq,
                    total_linea_gtq=total_linea_gtq,
                    iva_linea_gtq=iva_linea_gtq,
                    bien_o_servicio=it.get('bien_o_servicio', 'B')
                ))
            
            facturas_creadas.append(factura)
            
        except Exception as e:
            logger.error(f"Error en {file.filename}: {e}")
            rechazos.append(f"{file.filename}: {str(e)}")

    await db.commit()
    
    return {
        "cargadas": len(facturas_creadas),
        "rechazadas": rechazos,
        "requieren_revision_manual": requieren_revision_manual,  
    }


# ==========================================
# 2. Validar Hoja Electrónica
# ==========================================
@router.post("/validar-hoja-electronica", status_code=status.HTTP_200_OK)
async def validar_hoja_electronica(
    file: UploadFile = File(...),
    empresa_id: str | None = Query(None, description="ID de la empresa (opcional, usa X-Company-Id)"),  # ✅ Cambiado a opcional
    tenant_id: str | None = Query(None, description="ID del tenant (requerido para superadmin)"),
    scope: DataScope = Depends(get_data_scope),
    db: AsyncSession = Depends(get_public_db),
    empresa_from_header: Empresa | None = Depends(get_active_empresa)  # ✅ NUEVO
):
    await _set_schema_for_query(db, scope, tenant_id)
    
    # ✅ Usar empresa_id del header si no se pasó como query param
    empresa_id_final = empresa_id or (str(empresa_from_header.id) if empresa_from_header else None)
    
    if not empresa_id_final:
        raise HTTPException(400, detail="Debe especificar una empresa (query param o header X-Company-Id)")
    
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Solo se permiten archivos Excel (.xlsx/.xls)")
    
    try:
        content = await file.read()
        SHEET_NAME = "InformacionDTE-FEL"
        rows = []

        if file.filename.endswith('.xls'):
            book = xlrd.open_workbook(file_contents=content)
            if SHEET_NAME not in book.sheet_names():
                raise HTTPException(400, f"La hoja '{SHEET_NAME}' no existe en el archivo")
            ws = book.sheet_by_name(SHEET_NAME)
            for row_idx in range(1, ws.nrows):
                rows.append([ws.cell_value(row_idx, col) for col in range(ws.ncols)])
        else:
            wb = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
            if SHEET_NAME not in wb.sheetnames:
                wb.close()
                raise HTTPException(400, f"La hoja '{SHEET_NAME}' no existe en el archivo")
            ws = wb[SHEET_NAME]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()

        COL_AUTH = 1
        COL_ANULADO = 20
        COL_FECHA_ANULADO = 21
        COLS_IMPUESTOS = {
            22: 'petroleo', 23: 'turismo_hospedaje', 24: 'turismo_pasajes',
            25: 'timbre_prensa', 26: 'bomberos', 27: 'tasa_municipal',
            28: 'bebidas_alcoholicas', 29: 'tabaco', 30: 'cemento',
            31: 'bebidas_no_alcoholicas', 32: 'tarifa_portuaria'
        }

        pendientes = []
        actualizadas = 0
        impuestos_insertados = 0

        for row in rows:
            nueva_clasificacion = 'NORMAL'
            if not row or not row[COL_AUTH]:
                continue
            auth_raw = str(row[COL_AUTH]).strip()
            if not auth_raw:
                continue
            auth_clean = auth_raw.replace('.xml', '').replace('.XML', '').strip().upper()

            res = await db.execute(
                text("SELECT id, estado, clasificacion_gasto_sat FROM facturas_electronicas WHERE REPLACE(UPPER(xml_filename), '.XML', '') = :auth"),
                {"auth": auth_clean}
            )
            factura_row = res.first()

            if not factura_row:
                pendientes.append(auth_clean)
                continue
            else:
                await db.execute(
                    text("UPDATE facturas_electronicas SET validado = TRUE, fecha_validacion = CURRENT_TIMESTAMP WHERE id = :id"),
                    {"id": factura_row.id}
                )

            marca_anulado = str(row[COL_ANULADO]).strip() if row[COL_ANULADO] else "No"
            if marca_anulado.lower() == "si" and factura_row[1] != "Anulada":
                fecha_anul = row[COL_FECHA_ANULADO]
                if isinstance(fecha_anul, str):
                    fecha_anul = datetime.fromisoformat(fecha_anul.replace("Z", "+00:00"))
                
                await db.execute(
                    text("UPDATE facturas_electronicas SET estado = 'Anulada', fecha_anulacion = :fecha WHERE id = :id"),
                    {"fecha": fecha_anul, "id": factura_row.id}
                )
                actualizadas += 1

            for col_idx, tipo_codigo in COLS_IMPUESTOS.items():
                monto_val = row[col_idx]
                
                if monto_val and float(monto_val) > 0:
                    monto_dec = Decimal(str(monto_val))
                    
                    if tipo_codigo == 'COMBUSTIBLE':
                        nueva_clasificacion = 'COMBUSTIBLE'
                    elif tipo_codigo == 'HOTELES':
                        nueva_clasificacion = 'HOTEL_SERVICIOS'

            logger.info(f'Clasificación Gasto SAT: {factura_row.clasificacion_gasto_sat} -> {nueva_clasificacion}')
            if nueva_clasificacion != factura_row.clasificacion_gasto_sat:
                await db.execute(
                    text("UPDATE facturas_electronicas SET clasificacion_gasto_sat = :clasif WHERE id = :id"),
                    {"clasif": nueva_clasificacion, "id": factura_row.id}
                )

        await db.commit()

        if pendientes:
            return {"success": False, "mensaje": "Validación rechazada. Faltan cargar los siguientes XML antes de validar:", "pendientes": pendientes, "procesadas": actualizadas, "impuestos_registrados": impuestos_insertados}

        return {"success": True, "mensaje": "Validación exitosa. Estados e impuestos sincronizados.", "anulaciones_actualizadas": actualizadas, "impuestos_registrados": impuestos_insertados}

    except Exception as e:
        await db.rollback()
        raise HTTPException(500, f"Error procesando hoja electrónica: {str(e)}")


# ==========================================
# 3. Listar Facturas
# ==========================================
@router.get("/", response_model=List[FacturaOut])
async def listar_facturas(
    empresa_id: str | None = Query(None, description="Filtrar por empresa (opcional, usa X-Company-Id)"),  # ✅ Cambiado a opcional
    tenant_id: str | None = Query(None, description="ID del tenant (requerido para superadmin)"),
    scope: DataScope = Depends(get_data_scope),
    db: AsyncSession = Depends(get_public_db),
    empresa_from_header: Empresa | None = Depends(get_active_empresa)  # ✅ NUEVO
):
    await _set_schema_for_query(db, scope, tenant_id)
    
    # ✅ Usar empresa_id del header si no se pasó como query param
    empresa_id_final = empresa_id or (str(empresa_from_header.id) if empresa_from_header else None)
    
    stmt = (select(FacturaElectronica)
            .options(selectinload(FacturaElectronica.detalles))
            .order_by(FacturaElectronica.fecha_emision.desc()))
    
    if empresa_id_final:
        stmt = stmt.where(FacturaElectronica.empresa_id == empresa_id_final)

    result = await db.execute(stmt)
    return [FacturaOut.model_validate(f) for f in result.scalars().all()]


# ==========================================
# 4. Generar Partida desde Factura
# ==========================================
@router.post("/{factura_id}/generar-partida", response_model=PartidaOut, status_code=status.HTTP_201_CREATED)
async def generar_partida_desde_factura(
    factura_id: str,
    empresa_id: UUID | None = Query(None, description="ID de la empresa (opcional, usa X-Company-Id)"),  # ✅ Cambiado a opcional
    tenant_id: str | None = Query(None, description="ID del tenant (requerido para superadmin)"),
    scope: DataScope = Depends(get_data_scope),
    db: AsyncSession = Depends(get_public_db),
    empresa_from_header: Empresa | None = Depends(get_active_empresa)  # ✅ NUEVO
):
    await _set_schema_for_query(db, scope, tenant_id)
    
    # ✅ Usar empresa_id del header si no se pasó como query param
    empresa_id_final = empresa_id or (empresa_from_header.id if empresa_from_header else None)
    
    if not empresa_id_final:
        raise HTTPException(400, detail="Debe especificar una empresa (query param o header X-Company-Id)")
    
    stmt = select(FacturaElectronica).where(FacturaElectronica.id == factura_id, FacturaElectronica.empresa_id == empresa_id_final)
    result = await db.execute(stmt)
    factura = result.scalar_one_or_none()
    if not factura:
        raise HTTPException(status_code=404, detail="Factura no encontrada")

    cuenta_iva = await db.execute(select(CuentaContable).where(CuentaContable.empresa_id == empresa_id_final, CuentaContable.codigo == "1.1.4"))
    cuenta_iva = cuenta_iva.scalar_one_or_none()

    cuenta_gasto = await db.execute(select(CuentaContable).where(CuentaContable.empresa_id == empresa_id_final, CuentaContable.codigo == "5.1"))
    cuenta_gasto = cuenta_gasto.scalar_one_or_none()

    cuenta_proveedor = await db.execute(select(CuentaContable).where(CuentaContable.empresa_id == empresa_id_final, CuentaContable.codigo == "2.1.1"))
    cuenta_proveedor = cuenta_proveedor.scalar_one_or_none()

    if not all([cuenta_iva, cuenta_gasto, cuenta_proveedor]):
        raise HTTPException(status_code=400, detail="Faltan cuentas contables necesarias (1.1.4 IVA, 5.1 Costo de Ventas, 2.1.1 Proveedores)")

    detalles_partida = []
    if factura.total_gravado > 0:
        detalles_partida.append({"cuenta_id": cuenta_gasto.id, "tipo_movimiento": "debe", "monto": factura.total_gravado})
    if factura.total_iva > 0:
        detalles_partida.append({"cuenta_id": cuenta_iva.id, "tipo_movimiento": "debe", "monto": factura.total_iva})
    detalles_partida.append({"cuenta_id": cuenta_proveedor.id, "tipo_movimiento": "haber", "monto": factura.total})

    total_debe = sum(d["monto"] for d in detalles_partida if d["tipo_movimiento"] == "debe")
    total_haber = sum(d["monto"] for d in detalles_partida if d["tipo_movimiento"] == "haber")
    if total_debe != total_haber:
        raise HTTPException(status_code=400, detail="La partida no cuadra (debe ≠ haber)")

    partida = Partida(fecha=factura.fecha_emision.date(), descripcion=f"Factura {factura.serie} {factura.numero} - {factura.emisor_nombre}", numero_poliza=f"FEL-{factura.numero_autorizacion}")
    db.add(partida)
    await db.flush()

    for det in detalles_partida:
        db.add(DetallePartida(partida_id=partida.id, cuenta_id=det["cuenta_id"], tipo_movimiento=det["tipo_movimiento"], monto=det["monto"]))

    await db.commit()

    return PartidaOut(
        id=partida.id, numero=partida.numero, numero_poliza=partida.numero_poliza, fecha=partida.fecha,
        descripcion=partida.descripcion, empresa_nombre="", created_at=partida.created_at,
        detalles=[DetallePartidaOut(id=d.id, cuenta_id=d.cuenta_id, cuenta_codigo=d.cuenta.codigo if d.cuenta else "", cuenta_nombre=d.cuenta.nombre if d.cuenta else "", tipo_movimiento=d.tipo_movimiento, monto=d.monto) for d in partida.detalles]
    )


# ==========================================
# 5. Obtener Factura por ID
# ==========================================
@router.get("/{factura_id}", response_model=FacturaOut)
async def get_factura(
    factura_id: str,
    tenant_id: str | None = Query(None, description="ID del tenant (requerido para superadmin)"),
    scope: DataScope = Depends(get_data_scope),
    db: AsyncSession = Depends(get_public_db)
):
    await _set_schema_for_query(db, scope, tenant_id)
    try:
        uuid_obj = UUID(factura_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="ID de factura inválido. Formato esperado: UUID")

    stmt = select(FacturaElectronica).where(FacturaElectronica.id == uuid_obj).options(selectinload(FacturaElectronica.detalles))
    result = await db.execute(stmt)
    factura = result.scalar_one_or_none()
    if not factura:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    return factura


# ==========================================
# 6. Anular Factura
# ==========================================
@router.patch("/{factura_id}/anular", response_model=FacturaOut)
async def anular_factura(
    factura_id: str,
    tenant_id: str | None = Query(None, description="ID del tenant (requerido para superadmin)"),
    scope: DataScope = Depends(get_data_scope),
    db: AsyncSession = Depends(get_public_db)
):
    await _set_schema_for_query(db, scope, tenant_id)
    stmt = select(FacturaElectronica).where(FacturaElectronica.id == factura_id).options(selectinload(FacturaElectronica.detalles))
    result = await db.execute(stmt)
    factura = result.scalar_one_or_none()
    if not factura:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    if factura.estado == "Anulada":
        raise HTTPException(status_code=400, detail="La factura ya está anulada")

    factura.estado = "Anulada"
    await db.commit()
    await db.refresh(factura)
    return FacturaOut.model_validate(factura)


# ==========================================
# 7. Corrección Manual de Clasificación
# ==========================================
@router.patch("/{factura_id}/clasificacion", response_model=dict)
async def actualizar_clasificacion_gasto(
    factura_id: UUID,
    nueva_clasificacion: str,
    scope: DataScope = Depends(get_data_scope),
    db: AsyncSession = Depends(get_public_db)
):
    await _set_schema_for_query(db, scope)
    CATEGORIAS_VALIDAS = ['NORMAL', 'COMBUSTIBLE', 'ACTIVO_FIJO', 'MEDICAMENTO', 'PEQUENO_CONTRIBUYENTE', 'IMPORTACION']
    if nueva_clasificacion.upper() not in CATEGORIAS_VALIDAS:
        raise HTTPException(400, f"Categoría inválida. Use una de: {', '.join(CATEGORIAS_VALIDAS)}")

    await db.execute(
        text("UPDATE facturas_electronicas SET clasificacion_gasto_sat = :clasif WHERE id = :id"),
        {"clasif": nueva_clasificacion.upper(), "id": factura_id}
    )
    await db.commit()

    return {"success": True, "mensaje": "Clasificación actualizada correctamente"}