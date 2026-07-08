"""
Exportador genérico a Excel usando openpyxl.
Consume un ReportDefinition y produce un buffer BytesIO.
"""
import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models import ColumnType, ReportDefinition


class ExcelExporter:
    """Exporta cualquier ReportDefinition a Excel."""
    
    @staticmethod
    def export(report: ReportDefinition) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = report.title[:31]  # Excel limita a 31 chars
        
        # Estilos
        header_font = Font(bold=True, size=11, color=report.header_text_color)
        header_fill = PatternFill(
            start_color=report.header_background,
            end_color=report.header_background,
            fill_type="solid"
        )
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        title_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=12, color="FFFFFF")
        section_fill = PatternFill(start_color="#4472C4", end_color="#4472C4", fill_type="solid")
        
        current_row = 1
        
        # ===== HEADER DEL REPORTE =====
        ws.merge_cells(start_row=current_row, start_column=1, 
                      end_row=current_row, end_column=len(report.columns))
        cell = ws.cell(row=current_row, column=1, value=report.title)
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center')
        current_row += 1
        
        if report.subtitle:
            ws.merge_cells(start_row=current_row, start_column=1,
                          end_row=current_row, end_column=len(report.columns))
            cell = ws.cell(row=current_row, column=1, value=report.subtitle)
            cell.font = Font(italic=True, size=10)
            cell.alignment = Alignment(horizontal='center')
            current_row += 1
        
        if report.company_name or report.period:
            meta_parts = []
            if report.company_name:
                meta_parts.append(f"Empresa: {report.company_name}")
            if report.period:
                meta_parts.append(f"Período: {report.period}")
            ws.merge_cells(start_row=current_row, start_column=1,
                          end_row=current_row, end_column=len(report.columns))
            cell = ws.cell(row=current_row, column=1, value=" | ".join(meta_parts))
            cell.font = Font(size=9, color="666666")
            cell.alignment = Alignment(horizontal='center')
            current_row += 1
        
        current_row += 1  # Espacio
        
        # ===== SECCIONES =====
        for section in report.sections:
            # Título de sección (si existe)
            if section.title:
                ws.merge_cells(start_row=current_row, start_column=1,
                              end_row=current_row, end_column=len(report.columns))
                cell = ws.cell(row=current_row, column=1, value=section.title)
                cell.font = section_font
                cell.fill = section_fill
                cell.alignment = Alignment(horizontal='left')
                current_row += 1
            
            # Headers de columnas
            for col_idx, col in enumerate(report.columns, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=col.header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(
                    horizontal=col.alignment.value,
                    vertical='center'
                )
                cell.border = thin_border
                if col.width:
                    ws.column_dimensions[chr(64 + col_idx)].width = col.width
            current_row += 1
            
            # Filas de datos
            for row_idx, row in enumerate(section.rows):
                for col_idx, col in enumerate(report.columns, 1):
                    value = row.data.get(col.key, "")
                    # Formatear según tipo
                    value = ExcelExporter._format_value(value, col.type)
                    
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal=col.alignment.value)
                    cell.border = thin_border
                    
                    # Estilos de fila
                    if row.bold:
                        cell.font = Font(bold=True)
                    if row.italic:
                        cell.font = Font(italic=True)
                    if row.background_color:
                        cell.fill = PatternFill(
                            start_color=row.background_color,
                            end_color=row.background_color,
                            fill_type="solid"
                        )
                    elif report.alternate_row_color and row_idx % 2 == 1:
                        cell.fill = PatternFill(
                            start_color=report.alternate_row_color,
                            end_color=report.alternate_row_color,
                            fill_type="solid"
                        )
                    if row.text_color:
                        cell.font = Font(color=row.text_color)
                
                # Indentación (ej: plan de cuentas)
                if row.indent_level > 0:
                    first_cell = ws.cell(row=current_row, column=1)
                    first_cell.alignment = Alignment(
                        horizontal='left',
                        indent=row.indent_level * 2
                    )
                
                current_row += 1
            
            # Fila de totales de la sección
            if section.totals:
                for col_idx, col in enumerate(report.columns, 1):
                    value = section.totals.get(col.key, "")
                    value = ExcelExporter._format_value(value, col.type)
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.font = Font(bold=True, size=11)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal=col.alignment.value)
                current_row += 1
            
            current_row += 1  # Espacio entre secciones
        
        # ===== FOOTER =====
        if report.generated_at:
            current_row += 1
            ws.merge_cells(start_row=current_row, start_column=1,
                          end_row=current_row, end_column=len(report.columns))
            cell = ws.cell(row=current_row, column=1, 
                          value=f"Generado el {report.generated_at.strftime('%d/%m/%Y %H:%M')}")
            cell.font = Font(size=8, color="999999", italic=True)
        
        # Congelar paneles
        if report.freeze_panes:
            ws.freeze_panes = report.freeze_panes
        
        # Auto-filter
        if report.auto_filter and report.columns:
            # El auto-filter se aplica al header de la primera sección
            ws.auto_filter.ref = f"A2:{chr(64 + len(report.columns))}2"
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def _format_value(value: Any, col_type: ColumnType) -> Any:
        """Formatea un valor según el tipo de columna."""
        if value is None:
            return ""
        if col_type == ColumnType.CURRENCY:
            if isinstance(value, Decimal):
                return float(value)
            return value
        if col_type == ColumnType.NUMBER:
            if isinstance(value, Decimal):
                return float(value)
            return value
        if col_type == ColumnType.DATE:
            if isinstance(value, (date, datetime)):
                return value
            return value
        return value