import io
from datetime import date, datetime
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


# ---------------------------------------------------------------------------
def _serializar(valor):
    """Convierte objetos no serializables (Decimal, datetime) a tipos nativos."""
    if isinstance(valor, Decimal):
        return float(valor)
    if isinstance(valor, (datetime, date)):
        return valor.isoformat()
    return valor

# ================================================================
#  EXCEL
# ================================================================
def generar_balance_comprobacion_excel(datos: dict) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance de Comprobación"

    # Estilos
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = "Balance de Comprobación"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Encabezados
    headers = ["Código", "Cuenta", "Debe", "Haber", "Saldo"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Datos
    for i, fila in enumerate(datos.get("filas", []), 4):
        ws.cell(row=i, column=1, value=fila["codigo"]).border = thin_border
        ws.cell(row=i, column=2, value=fila["nombre"]).border = thin_border
        ws.cell(row=i, column=3, value=float(fila["sum_debe"])).border = thin_border
        ws.cell(row=i, column=4, value=float(fila["sum_haber"])).border = thin_border
        ws.cell(row=i, column=5, value=float(fila["saldo"])).border = thin_border

    # Ancho columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generar_estado_resultados_excel(datos: dict) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estado de Resultados"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:C1')
    ws['A1'] = "Estado de Resultados"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    row = 3
    # Ingresos
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=1, value="INGRESOS").font = Font(bold=True, size=12)
    row += 1
    for col, h in enumerate(["Código", "Cuenta", "Monto"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    row += 1
    for ing in datos.get("ingresos", []):
        ws.cell(row=row, column=1, value=ing["codigo"]).border = thin_border
        ws.cell(row=row, column=2, value=ing["nombre"]).border = thin_border
        ws.cell(row=row, column=3, value=float(ing["saldo"])).border = thin_border
        row += 1
    # Total ingresos
    ws.cell(row=row, column=1, value="Total Ingresos").font = Font(bold=True)
    ws.cell(row=row, column=3, value=float(datos["total_ingresos"])).font = Font(bold=True)
    row += 2

    # Gastos
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=1, value="GASTOS").font = Font(bold=True, size=12)
    row += 1
    for col, h in enumerate(["Código", "Cuenta", "Monto"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    row += 1
    for gas in datos.get("gastos", []):
        ws.cell(row=row, column=1, value=gas["codigo"]).border = thin_border
        ws.cell(row=row, column=2, value=gas["nombre"]).border = thin_border
        ws.cell(row=row, column=3, value=float(gas["saldo"])).border = thin_border
        row += 1
    # Total gastos
    ws.cell(row=row, column=1, value="Total Gastos").font = Font(bold=True)
    ws.cell(row=row, column=3, value=float(datos["total_gastos"])).font = Font(bold=True)
    row += 2

    # Utilidad neta
    ws.cell(row=row, column=1, value="UTILIDAD NETA").font = Font(bold=True, size=12, color="008000")
    ws.cell(row=row, column=3, value=float(datos["utilidad_neta"])).font = Font(bold=True, size=12, color="008000")

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generar_balance_general_excel(datos: dict) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance General"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:C1')
    ws['A1'] = "Balance General"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    row = 3
    for seccion, titulo in [("activos", "ACTIVOS"), ("pasivos", "PASIVOS"), ("patrimonio", "PATRIMONIO")]:
        ws.merge_cells(f'A{row}:C{row}')
        ws.cell(row=row, column=1, value=titulo).font = Font(bold=True, size=12)
        row += 1
        for col, h in enumerate(["Código", "Cuenta", "Saldo"], 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
        row += 1
        for f in datos.get(seccion, []):
            ws.cell(row=row, column=1, value=f["codigo"]).border = thin_border
            ws.cell(row=row, column=2, value=f["nombre"]).border = thin_border
            ws.cell(row=row, column=3, value=float(f["saldo"])).border = thin_border
            row += 1
        # Total de la sección
        ws.cell(row=row, column=1, value=f"Total {titulo}").font = Font(bold=True)
        ws.cell(row=row, column=3, value=float(datos[f"total_{seccion}"])).font = Font(bold=True)
        row += 2

    ws.cell(row=row, column=1, value="UTILIDAD DEL EJERCICIO").font = Font(bold=True, size=12, color="008000")
    ws.cell(row=row, column=3, value=float(datos["utilidad_ejercicio"])).font = Font(bold=True, size=12, color="008000")

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ================================================================
#  PDF
# ================================================================
def generar_balance_comprobacion_pdf(datos: dict) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph("Balance de Comprobación", styles['Title']))
    elements.append(Spacer(1, 0.25*inch))

    # Tabla
    data = [["Código", "Cuenta", "Debe", "Haber", "Saldo"]]
    for f in datos.get("filas", []):
        data.append([
            f["codigo"],
            f["nombre"],
            f"{float(f['sum_debe']):,.2f}",
            f"{float(f['sum_haber']):,.2f}",
            f"{float(f['saldo']):,.2f}"
        ])

    table = Table(data, colWidths=[1.0*inch, 2.8*inch, 1.0*inch, 1.0*inch, 1.0*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4472C4")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (2,0), (-1,-1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generar_estado_resultados_pdf(datos: dict) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph("Estado de Resultados", styles['Title']))
    elements.append(Spacer(1, 0.25*inch))

    # Ingresos
    elements.append(Paragraph("INGRESOS", styles['Heading2']))
    data = [["Código", "Cuenta", "Monto"]]
    for ing in datos.get("ingresos", []):
        data.append([ing["codigo"], ing["nombre"], f"{float(ing['saldo']):,.2f}"])
    data.append(["", "Total Ingresos", f"{float(datos['total_ingresos']):,.2f}"])
    table = Table(data, colWidths=[1.0*inch, 3.0*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4472C4")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (2,0), (-1,-1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.25*inch))

    # Gastos
    elements.append(Paragraph("GASTOS", styles['Heading2']))
    data = [["Código", "Cuenta", "Monto"]]
    for gas in datos.get("gastos", []):
        data.append([gas["codigo"], gas["nombre"], f"{float(gas['saldo']):,.2f}"])
    data.append(["", "Total Gastos", f"{float(datos['total_gastos']):,.2f}"])
    table = Table(data, colWidths=[1.0*inch, 3.0*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4472C4")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (2,0), (-1,-1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.25*inch))

    # Utilidad
    elements.append(Paragraph(f"<b>UTILIDAD NETA: {float(datos['utilidad_neta']):,.2f}</b>", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generar_balance_general_pdf(datos: dict) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph("Balance General", styles['Title']))
    elements.append(Spacer(1, 0.25*inch))

    for seccion, titulo in [("activos", "ACTIVOS"), ("pasivos", "PASIVOS"), ("patrimonio", "PATRIMONIO")]:
        elements.append(Paragraph(titulo, styles['Heading2']))
        data = [["Código", "Cuenta", "Saldo"]]
        for f in datos.get(seccion, []):
            data.append([f["codigo"], f["nombre"], f"{float(f['saldo']):,.2f}"])
        data.append(["", f"Total {titulo}", f"{float(datos[f'total_{seccion}']):,.2f}"])
        table = Table(data, colWidths=[1.0*inch, 3.0*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4472C4")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (2,0), (-1,-1), 'RIGHT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.15*inch))

    elements.append(Paragraph(f"<b>UTILIDAD DEL EJERCICIO: {float(datos['utilidad_ejercicio']):,.2f}</b>", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return buffer