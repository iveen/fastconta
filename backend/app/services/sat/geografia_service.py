"""Servicio para gestión de Geografía (Departamentos y Municipios)"""

from io import BytesIO
from uuid import UUID

from app.models.global_models import Departamento, Municipio
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

# Columnas para exportación
COLUMNAS_DEPTO_EXPORT = [
    {"key": "codigo_iso", "header": "Código ISO", "width": 12},
    {"key": "nombre", "header": "Departamento", "width": 30},
    {"key": "total_municipios", "header": "Total Municipios", "width": 18},
]

COLUMNAS_MUNICIPIO_EXPORT = [
    {"key": "codigo_iso", "header": "Código ISO", "width": 12},
    {"key": "nombre", "header": "Municipio", "width": 30},
    {"key": "departamento_codigo", "header": "Código Depto", "width": 12},
    {"key": "departamento_nombre", "header": "Departamento", "width": 30},
]

COLUMNAS_IMPORT_DEPTO = ["codigo_iso", "nombre"]
COLUMNAS_IMPORT_MUNICIPIO = ["codigo_iso", "nombre", "departamento_codigo"]


class GeografiaService:
    def __init__(self, db: AsyncSession):
        self.db = db

    # ============================================================
    # DEPARTAMENTOS - QUERIES
    # ============================================================
    async def obtener_departamentos(
        self,
        search: str | None = None,
        skip: int = 0,
        limit: int = 50,
    ) -> tuple[list[Departamento], int]:
        """Lista departamentos con conteo de municipios"""
        query = (
            select(Departamento)
            .options(selectinload(Departamento.municipios))
        )

        if search:
            search_pattern = f"%{search}%"
            query = query.where(
                (Departamento.nombre.ilike(search_pattern))
                | (Departamento.codigo_iso.ilike(search_pattern))
            )

        count_query = select(func.count()).select_from(query.subquery())
        total = (await self.db.execute(count_query)).scalar() or 0

        query = query.order_by(Departamento.nombre).offset(skip).limit(limit)
        result = await self.db.execute(query)
        return list(result.scalars().all()), total

    async def obtener_departamento_por_id(
        self, departamento_id: UUID
    ) -> Departamento | None:
        """Obtiene un departamento con sus municipios"""
        query = (
            select(Departamento)
            .where(Departamento.id == departamento_id)
            .options(selectinload(Departamento.municipios))
        )
        result = await self.db.execute(query)
        return result.scalars().first()

    async def obtener_departamento_por_codigo(
        self, codigo_iso: str
    ) -> Departamento | None:
        """Obtiene un departamento por código ISO"""
        query = select(Departamento).where(Departamento.codigo_iso == codigo_iso)
        result = await self.db.execute(query)
        return result.scalars().first()

    async def obtener_todos_departamentos(self) -> list[Departamento]:
        """Obtiene todos los departamentos (para dropdowns)"""
        query = (
            select(Departamento)
            .order_by(Departamento.nombre)
        )
        result = await self.db.execute(query)
        return list(result.scalars().all())

    # ============================================================
    # DEPARTAMENTOS - CRUD
    # ============================================================
    async def crear_departamento(self, data: dict) -> Departamento:
        """Crea un nuevo departamento"""
        existente = await self.obtener_departamento_por_codigo(data["codigo_iso"])
        if existente is not None:
            raise ValueError(
                f"Ya existe un departamento con código '{data['codigo_iso']}'"
            )

        depto = Departamento(**data)
        self.db.add(depto)
        await self.db.flush()
        await self.db.refresh(depto)
        return depto

    async def actualizar_departamento(
        self, departamento_id: UUID, data: dict
    ) -> Departamento | None:
        """Actualiza un departamento"""
        depto = await self.obtener_departamento_por_id(departamento_id)
        if depto is None:
            return None

        for key, value in data.items():
            if hasattr(depto, key):
                setattr(depto, key, value)

        await self.db.flush()
        await self.db.refresh(depto)
        return depto

    async def eliminar_departamento(self, departamento_id: UUID) -> bool:
        """Elimina un departamento (hard delete con cascade a municipios)"""
        depto = await self.obtener_departamento_por_id(departamento_id)
        if depto is None:
            return False

        await self.db.delete(depto)
        await self.db.flush()
        return True

    # ============================================================
    # MUNICIPIOS - QUERIES
    # ============================================================
    async def obtener_municipios(
        self,
        departamento_id: UUID | None = None,
        search: str | None = None,
        skip: int = 0,
        limit: int = 100,
    ) -> tuple[list[Municipio], int]:
        """Lista municipios con filtros"""
        query = (
            select(Municipio)
            .options(selectinload(Municipio.departamento))
        )

        if departamento_id is not None:
            query = query.where(Municipio.departamento_id == departamento_id)
        if search:
            search_pattern = f"%{search}%"
            query = query.where(
                (Municipio.nombre.ilike(search_pattern))
                | (Municipio.codigo_iso.ilike(search_pattern))
            )

        count_query = select(func.count()).select_from(query.subquery())
        total = (await self.db.execute(count_query)).scalar() or 0

        query = query.order_by(Municipio.codigo_iso).offset(skip).limit(limit)
        result = await self.db.execute(query)
        return list(result.scalars().all()), total

    async def obtener_municipio_por_id(self, municipio_id: UUID) -> Municipio | None:
        """Obtiene un municipio con su departamento"""
        query = (
            select(Municipio)
            .where(Municipio.id == municipio_id)
            .options(selectinload(Municipio.departamento))
        )
        result = await self.db.execute(query)
        return result.scalars().first()

    async def obtener_municipio_por_codigo(
        self, codigo_iso: str
    ) -> Municipio | None:
        """Obtiene un municipio por código ISO"""
        query = select(Municipio).where(Municipio.codigo_iso == codigo_iso)
        result = await self.db.execute(query)
        return result.scalars().first()

    # ============================================================
    # MUNICIPIOS - CRUD
    # ============================================================
    async def crear_municipio(self, data: dict) -> Municipio:
        """Crea un nuevo municipio"""
        # Validar departamento
        depto = await self.obtener_departamento_por_id(data["departamento_id"])
        if depto is None:
            raise ValueError("Departamento no encontrado")

        # Validar código único
        existente = await self.obtener_municipio_por_codigo(data["codigo_iso"])
        if existente is not None:
            raise ValueError(
                f"Ya existe un municipio con código '{data['codigo_iso']}'"
            )

        municipio = Municipio(**data)
        self.db.add(municipio)
        await self.db.flush()
        await self.db.refresh(municipio)
        return municipio

    async def actualizar_municipio(
        self, municipio_id: UUID, data: dict
    ) -> Municipio | None:
        """Actualiza un municipio"""
        municipio = await self.obtener_municipio_por_id(municipio_id)
        if municipio is None:
            return None

        # Validar departamento si cambia
        if "departamento_id" in data:
            depto = await self.obtener_departamento_por_id(data["departamento_id"])
            if depto is None:
                raise ValueError("Departamento no encontrado")

        for key, value in data.items():
            if hasattr(municipio, key):
                setattr(municipio, key, value)

        await self.db.flush()
        await self.db.refresh(municipio)
        return municipio

    async def eliminar_municipio(self, municipio_id: UUID) -> bool:
        """Elimina un municipio"""
        municipio = await self.obtener_municipio_por_id(municipio_id)
        if municipio is None:
            return False

        await self.db.delete(municipio)
        await self.db.flush()
        return True

    # ============================================================
    # IMPORT/EXPORT
    # ============================================================
    async def exportar_excel(self) -> BytesIO:
        """Exporta departamentos y municipios a Excel (2 hojas)"""
        # Hoja 1: Departamentos
        deptos = await self.obtener_todos_departamentos()
        datos_deptos = [
            {
                "codigo_iso": d.codigo_iso,
                "nombre": d.nombre,
                "total_municipios": len(d.municipios),
            }
            for d in deptos
        ]

        # Hoja 2: Municipios
        municipios, _ = await self.obtener_municipios(limit=500)
        datos_municipios = [
            {
                "codigo_iso": m.codigo_iso,
                "nombre": m.nombre,
                "departamento_codigo": m.departamento.codigo_iso if m.departamento else "",
                "departamento_nombre": m.departamento.nombre if m.departamento else "",
            }
            for m in municipios
        ]

        # Crear workbook con 2 hojas
        from openpyxl import Workbook

        wb = Workbook()

        # Hoja Departamentos
        ws1 = wb.active
        ws1.title = "Departamentos"
        self._escribir_hoja(ws1, COLUMNAS_DEPTO_EXPORT, datos_deptos, "Departamentos de Guatemala")

        # Hoja Municipios
        ws2 = wb.create_sheet("Municipios")
        self._escribir_hoja(ws2, COLUMNAS_MUNICIPIO_EXPORT, datos_municipios, "Municipios de Guatemala")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _escribir_hoja(self, ws, columnas, datos, titulo):
        """Helper para escribir una hoja Excel"""
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        fila_actual = 1

        # Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
        celda_titulo = ws.cell(row=1, column=1, value=titulo)
        celda_titulo.font = Font(bold=True, size=14)
        celda_titulo.alignment = Alignment(horizontal="center")
        fila_actual = 3

        # Encabezados
        for col_idx, columna in enumerate(columnas, 1):
            celda = ws.cell(row=fila_actual, column=col_idx, value=columna["header"])
            celda.font = header_font
            celda.fill = header_fill
            celda.alignment = header_alignment
            if "width" in columna:
                ws.column_dimensions[get_column_letter(col_idx)].width = columna["width"]

        fila_actual += 1

        # Datos
        for fila_datos in datos:
            for col_idx, columna in enumerate(columnas, 1):
                valor = fila_datos.get(columna["key"], "")
                celda = ws.cell(row=fila_actual, column=col_idx, value=valor)
                celda.alignment = Alignment(vertical="center")
            fila_actual += 1

    async def importar_excel(
        self,
        archivo_bytes: bytes,
        sobrescribir: bool = False,
    ) -> dict:
        """
        Importa departamentos y municipios desde Excel (2 hojas).
        
        Returns:
            Dict con estadísticas de importación
        """
        from io import BytesIO

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(archivo_bytes), read_only=True)

        deptos_creados = 0
        deptos_actualizados = 0
        municipios_creados = 0
        municipios_actualizados = 0
        omitidos = 0
        errores: list[str] = []

        # ============================================================
        # Hoja 1: Departamentos
        # ============================================================
        if "Departamentos" in wb.sheetnames:
            ws = wb["Departamentos"]
            filas = self._leer_hoja(ws, COLUMNAS_IMPORT_DEPTO)

            for idx, fila in enumerate(filas, 2):
                try:
                    codigo_iso = str(fila.get("codigo_iso", "")).strip().upper()
                    nombre = str(fila.get("nombre", "")).strip()

                    if not codigo_iso or not nombre:
                        errores.append(f"Depto fila {idx}: Código y nombre obligatorios")
                        continue

                    if len(codigo_iso) != 2:
                        errores.append(f"Depto fila {idx}: Código debe tener 2 caracteres")
                        continue

                    if len(nombre) > 100:
                        errores.append(f"Depto fila {idx}: Nombre excede 100 caracteres")
                        continue

                    existente = await self.obtener_departamento_por_codigo(codigo_iso)

                    if existente is not None:
                        if sobrescribir:
                            existente.nombre = nombre
                            deptos_actualizados += 1
                        else:
                            omitidos += 1
                    else:
                        depto = Departamento(codigo_iso=codigo_iso, nombre=nombre)
                        self.db.add(depto)
                        deptos_creados += 1

                except Exception as e:
                    errores.append(f"Depto fila {idx}: {str(e)}")

        # ============================================================
        # Hoja 2: Municipios
        # ============================================================
        if "Municipios" in wb.sheetnames:
            ws = wb["Municipios"]
            filas = self._leer_hoja(ws, COLUMNAS_IMPORT_MUNICIPIO)

            for idx, fila in enumerate(filas, 2):
                try:
                    codigo_iso = str(fila.get("codigo_iso", "")).strip()
                    nombre = str(fila.get("nombre", "")).strip()
                    depto_codigo = str(fila.get("departamento_codigo", "")).strip().upper()

                    if not codigo_iso or not nombre or not depto_codigo:
                        errores.append(f"Muni fila {idx}: Todos los campos son obligatorios")
                        continue

                    if len(codigo_iso) != 4:
                        errores.append(f"Muni fila {idx}: Código debe tener 4 caracteres")
                        continue

                    if len(nombre) > 100:
                        errores.append(f"Muni fila {idx}: Nombre excede 100 caracteres")
                        continue

                    # Buscar departamento
                    depto = await self.obtener_departamento_por_codigo(depto_codigo)
                    if depto is None:
                        errores.append(f"Muni fila {idx}: Departamento '{depto_codigo}' no encontrado")
                        continue

                    # Buscar municipio existente
                    existente = await self.obtener_municipio_por_codigo(codigo_iso)

                    if existente is not None:
                        if sobrescribir:
                            existente.nombre = nombre
                            existente.departamento_id = depto.id
                            municipios_actualizados += 1
                        else:
                            omitidos += 1
                    else:
                        municipio = Municipio(
                            codigo_iso=codigo_iso,
                            nombre=nombre,
                            departamento_id=depto.id,
                        )
                        self.db.add(municipio)
                        municipios_creados += 1

                except Exception as e:
                    errores.append(f"Muni fila {idx}: {str(e)}")

        wb.close()
        await self.db.flush()

        return {
            "departamentos_creados": deptos_creados,
            "departamentos_actualizados": deptos_actualizados,
            "municipios_creados": municipios_creados,
            "municipios_actualizados": municipios_actualizados,
            "omitidos": omitidos,
            "errores": errores,
        }

    def _leer_hoja(self, ws, columnas_requeridas):
        """Helper para leer una hoja Excel"""
        encabezados = []
        for cell in ws[1]:
            if cell.value is None:
                break
            encabezados.append(str(cell.value).strip())

        # Saltar fila de título si existe
        if len(encabezados) == 0 or encabezados[0] == "":
            # Buscar la fila de encabezados reales
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                if row and row[0] and str(row[0]).strip() in columnas_requeridas:
                    encabezados = [str(v).strip() if v else "" for v in row]
                    break

        # Leer datos (asumiendo que los encabezados están en fila 1 o 3)
        datos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue

            fila_dict = {}
            for idx, valor in enumerate(row):
                if idx < len(encabezados):
                    fila_dict[encabezados[idx]] = valor

            if fila_dict:
                datos.append(fila_dict)

        return datos