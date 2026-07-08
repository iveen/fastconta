"""Service para Tipos DTE"""
import io
from uuid import UUID

from app.models.global_models import TipoDTE
from app.schemas.catalogos.tipo_dte import (
    TipoDTEImportResult,
)
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession


class TipoDTEService:
    def __init__(self, db: AsyncSession):
        self.db = db

    # ---------------------------------------------------------------
    # LISTAR CON PAGINACIÓN Y FILTROS
    # ---------------------------------------------------------------
    async def obtener_todos(
        self,
        activo: bool | None = None,
        es_factura: bool | None = None,
        skip: int = 0,
        limit: int = 50,
    ) -> tuple[list[TipoDTE], int]:
        query = select(TipoDTE)
        count_query = select(func.count()).select_from(TipoDTE)

        if activo is not None:
            query = query.where(TipoDTE.activo == activo)
            count_query = count_query.where(TipoDTE.activo == activo)

        if es_factura is not None:
            query = query.where(TipoDTE.es_factura == es_factura)
            count_query = count_query.where(TipoDTE.es_factura == es_factura)

        total = (await self.db.execute(count_query)).scalar_one()
        query = query.order_by(TipoDTE.codigo).offset(skip).limit(limit)
        result = await self.db.execute(query)
        return list(result.scalars().all()), total

    # ---------------------------------------------------------------
    # LISTAR ACTIVOS (para dropdowns)
    # ---------------------------------------------------------------
    async def obtener_todos_activos(self) -> list[TipoDTE]:
        query = (
            select(TipoDTE)
            .where(TipoDTE.activo.is_(True))
            .order_by(TipoDTE.codigo)
        )
        result = await self.db.execute(query)
        return list(result.scalars().all())

    # ---------------------------------------------------------------
    # OBTENER POR ID
    # ---------------------------------------------------------------
    async def obtener_por_id(self, dte_id: UUID) -> TipoDTE | None:
        query = select(TipoDTE).where(TipoDTE.id == dte_id)
        result = await self.db.execute(query)
        return result.scalar_one_or_none()

    # ---------------------------------------------------------------
    # CREAR
    # ---------------------------------------------------------------
    async def crear(self, data: dict) -> TipoDTE:
        existente = await self.db.execute(
            select(TipoDTE).where(TipoDTE.codigo == data["codigo"])
        )
        if existente.scalar_one_or_none():
            raise ValueError(f"Ya existe un Tipo DTE con código '{data['codigo']}'")

        dte = TipoDTE(**data)
        self.db.add(dte)
        await self.db.commit()
        await self.db.refresh(dte)
        return dte

    # ---------------------------------------------------------------
    # ACTUALIZAR
    # ---------------------------------------------------------------
    async def actualizar(self, dte_id: UUID, data: dict) -> TipoDTE | None:
        dte = await self.obtener_por_id(dte_id)
        if not dte:
            return None

        for campo, valor in data.items():
            setattr(dte, campo, valor)

        await self.db.commit()
        await self.db.refresh(dte)
        return dte

    # ---------------------------------------------------------------
    # ELIMINAR (soft delete)
    # ---------------------------------------------------------------
    async def eliminar(self, dte_id: UUID) -> bool:
        dte = await self.obtener_por_id(dte_id)
        if not dte:
            return False
        dte.activo = False
        await self.db.commit()
        return True

    # ---------------------------------------------------------------
    # EXPORTAR A EXCEL
    # ---------------------------------------------------------------
    async def exportar_excel(self) -> io.BytesIO:
        try:
            import openpyxl
        except ImportError:
            raise ValueError("openpyxl no está instalado")

        query = select(TipoDTE).order_by(TipoDTE.codigo)
        result = await self.db.execute(query)
        dtes = result.scalars().all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tipos DTE"
        ws.append(["Código", "Descripción", "Requiere Complemento", "Es Factura", "Activo"])

        for dte in dtes:
            ws.append([dte.codigo, dte.descripcion, dte.requiere_complemento, dte.es_factura, dte.activo])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ---------------------------------------------------------------
    # IMPORTAR DESDE EXCEL
    # ---------------------------------------------------------------
    async def importar_excel(
        self, archivo_bytes: bytes, sobrescribir: bool = False
    ) -> TipoDTEImportResult:
        try:
            import openpyxl
        except ImportError:
            raise ValueError("openpyxl no está instalado")

        wb = openpyxl.load_workbook(io.BytesIO(archivo_bytes))
        ws = wb.active

        resultado = TipoDTEImportResult()
        filas = list(ws.iter_rows(min_row=2, values_only=True))  # saltar encabezado

        for fila in filas:
            try:
                codigo, descripcion, req_comp, es_factura = fila[:4]
                if not codigo or not descripcion:
                    resultado.omitidos += 1
                    continue

                existente = await self.db.execute(
                    select(TipoDTE).where(TipoDTE.codigo == codigo)
                )
                dte = existente.scalar_one_or_none()

                if dte:
                    if sobrescribir:
                        dte.descripcion = descripcion
                        dte.requiere_complemento = bool(req_comp)
                        dte.es_factura = bool(es_factura)
                        resultado.actualizados += 1
                    else:
                        resultado.omitidos += 1
                else:
                    nuevo = TipoDTE(
                        codigo=codigo,
                        descripcion=descripcion,
                        requiere_complemento=bool(req_comp),
                        es_factura=bool(es_factura),
                    )
                    self.db.add(nuevo)
                    resultado.creados += 1
            except Exception as e:
                resultado.errores.append(f"Fila {fila}: {str(e)}")

        await self.db.commit()
        return resultado