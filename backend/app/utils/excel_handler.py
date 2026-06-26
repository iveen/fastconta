"""Utilidad reutilizable para import/export Excel"""

from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelHandler:
    """Handler reutilizable para operaciones Excel"""

    @staticmethod
    def exportar_a_excel(
        datos: list[dict],
        columnas: list[dict],
        nombre_hoja: str = "Datos",
        titulo: str | None = None,
    ) -> BytesIO:
        """
        Exporta datos a Excel.
        
        Args:
            datos: Lista de diccionarios con los datos
            columnas: Lista de dicts con {'key': 'campo', 'header': 'Encabezado', 'width': 15}
            nombre_hoja: Nombre de la hoja
            titulo: Título opcional en la primera fila
        
        Returns:
            BytesIO con el archivo Excel
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = nombre_hoja

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        fila_actual = 1

        # Título
        if titulo:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
            celda_titulo = ws.cell(row=1, column=1, value=titulo)
            celda_titulo.font = Font(bold=True, size=14)
            celda_titulo.alignment = Alignment(horizontal="center")
            fila_actual = 2

        # Encabezados
        for col_idx, columna in enumerate(columnas, 1):
            celda = ws.cell(row=fila_actual, column=col_idx, value=columna["header"])
            celda.font = header_font
            celda.fill = header_fill
            celda.alignment = header_alignment
            if "width" in columna:
                ws.column_dimensions[get_column_letter(col_idx)].width = columna["width"]

        fila_actual += 1

        # Datos
        for fila_datos in datos:
            for col_idx, columna in enumerate(columnas, 1):
                valor = fila_datos.get(columna["key"], "")
                celda = ws.cell(row=fila_actual, column=col_idx, value=valor)
                celda.alignment = Alignment(vertical="center")
            fila_actual += 1

        # Auto-ajustar ancho si no se especificó
        for col_idx in range(1, len(columnas) + 1):
            letra = get_column_letter(col_idx)
            if ws.column_dimensions[letra].width is None:
                ws.column_dimensions[letra].width = 15

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def importar_desde_excel(
        archivo_bytes: bytes,
        columnas_requeridas: list[str],
        nombre_hoja: str | None = None,
        fila_inicio: int = 1,
    ) -> list[dict[str, Any]]:
        """
        Importa datos desde Excel.
        
        Args:
            archivo_bytes: Contenido del archivo Excel
            columnas_requeridas: Lista de nombres de columnas requeridas
            nombre_hoja: Nombre de la hoja (None = primera hoja)
            fila_inicio: Fila donde empiezan los datos (1-based)
        
        Returns:
            Lista de diccionarios con los datos
        """
        wb = openpyxl.load_workbook(BytesIO(archivo_bytes), read_only=True)
        ws = wb[nombre_hoja] if nombre_hoja else wb.active

        # Leer encabezados
        encabezados = []
        for cell in ws[1]:
            if cell.value is None:
                break
            encabezados.append(str(cell.value).strip())

        # Validar columnas requeridas
        columnas_faltantes = [c for c in columnas_requeridas if c not in encabezados]
        if columnas_faltantes:
            raise ValueError(f"Columnas faltantes en el Excel: {columnas_faltantes}")

        # Leer datos
        datos = []
        for row in ws.iter_rows(min_row=fila_inicio + 1, values_only=True):
            if all(v is None for v in row):
                continue  # Saltar filas vacías
            
            fila_dict = {}
            for idx, valor in enumerate(row):
                if idx < len(encabezados):
                    fila_dict[encabezados[idx]] = valor
            
            if fila_dict:
                datos.append(fila_dict)

        wb.close()
        return datos